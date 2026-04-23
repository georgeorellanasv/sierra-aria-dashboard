"""
Build a structured issue-log Excel from the Sonnet classifications.

Output: reports/issue_log.xlsx with columns matching the user's template:
  JOURNEY | ISSUE | DESCRIPTION | # IMPACTED | PROPOSED OUTCOME
  | METRIC (VOLUME) | RAISED BY | DATE LOGGED | SEVERITY | REFERENCE | STATUS

Method:
  1. Load all high/critical classifications (most signal-rich).
  2. Send the pain-point corpus to Sonnet and ask for N thematic issues with
     title, journey, severity, representative session_ids, and a detailed
     PROPOSED OUTCOME in IF/THEN flow syntax.
  3. For each issue, pull literal quotes from the referenced session's
     transcript to populate DESCRIPTION.
  4. Write to Excel with formatting.

Cost: ~$0.30-0.50 Sonnet usage for the clustering+synthesis call.
"""
from __future__ import annotations

import json
import logging
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from anthropic import Anthropic
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src import config

logger = logging.getLogger("issue_log")
logging.basicConfig(level=logging.INFO,
                    format="%(asctime)s  %(levelname)-7s %(name)s  %(message)s")

MODEL = "claude-sonnet-4-6"
OUT_PATH = Path(__file__).resolve().parent.parent / "reports" / "issue_log.xlsx"

# Map classifier categories to journey labels (how they should appear in the log).
CATEGORY_TO_JOURNEY = {
    "transaction_status":  "Order Status",
    "cancel_transaction":  "Cancel Order",
    "authentication":      "Authentication",
    "refund":              "Refund",
    "transfer_to_human":   "Escalation",
    "technical_issue":     "Technical",
    "complaint":           "Complaint",
    "general_info":        "General Info",
    "greeting_drop":       "Greeting",
    "other":               "Other",
}

CLUSTER_PROMPT = """You are analyzing Sierra voice-agent session reviews to \
produce a consolidated issue log for engineering review.

Input: a JSON array of per-session classifications. Each entry has:
  - session_id
  - category     (transaction_status, authentication, etc.)
  - severity     (critical | high | medium | low)
  - pain_points  (list of short descriptions of what went wrong)
  - suggestion   (verbose recommendation from the per-session review)
  - transcript_excerpt (a few real lines from the call for grounding)

Your job: cluster the pain points into **12 to 20 distinct, actionable \
issues**. Each issue must:

  1. Group pain points that share the same root cause (not just the same topic).
  2. Cover the highest-volume + highest-severity problems first.
  3. Reference **at least one session_id** as evidence (use up to 3).
  4. Have a PROPOSED_OUTCOME in IF/THEN flow syntax — concrete, not generic.

Output strict JSON only (no prose, no markdown fences), matching:

{
  "issues": [
    {
      "journey":           "Order Status",
      "issue_title":       "Short title, e.g. 'Agent loops on order number after STT mis-parse'",
      "description":       "Plain-English summary of what goes wrong, ~2-4 sentences. Reference ONE literal caller quote if present in the transcript excerpts.",
      "impacted_count":    <int>  // number of input sessions with this issue
      "proposed_outcome":  "Multi-line fix using IF/THEN syntax. Name exact journey blocks and tools. Example:\\n\\nIN 'Intents Where User Needs to Authenticate':\\n  IF tool:CustomerByOrderNumber fails 2x\\n  >> call tool:CustomerByTelephone using caller ANI\\n  IF that also fails\\n  >> tool:CreateZendeskTicket(reason='unable_to_authenticate')\\n  >> inform caller a human will follow up, end call.",
      "severity":          "Critical | High | Medium | Low",
      "reference_sessions": ["audit-01...", "audit-02..."]
    }
    ...
  ]
}

Rules:
- journey must be one of: Order Status, Cancel Order, Authentication, Refund, \
Escalation, Technical, Complaint, General Info, Greeting, Other
- severity mirrors the worst severity among the underlying sessions
- reference_sessions is a subset of the input session_ids (do not invent IDs)
- impacted_count is the number of input sessions this issue represents
- issues must be ordered by (severity critical>high>medium>low, then impacted_count desc)
"""


# ---------------------------------------------------------------------------
# Data gathering
# ---------------------------------------------------------------------------

def gather_classifications(conn: sqlite3.Connection) -> list[dict]:
    """Return one dict per high/critical classified session, with pain points
    + suggestion + a short transcript excerpt."""
    rows = conn.execute("""
        SELECT c.session_id, c.category, c.severity, c.pain_points_json,
               c.suggestion, c.related_journey_blocks,
               s.duration_seconds, s.first_user_message
        FROM classifications c
        JOIN sessions s ON s.id = c.session_id
        WHERE c.severity IN ('critical', 'high', 'medium')
        ORDER BY
          CASE c.severity
            WHEN 'critical' THEN 0 WHEN 'high' THEN 1 WHEN 'medium' THEN 2 ELSE 3 END,
          s.duration_seconds DESC
    """).fetchall()

    out: list[dict] = []
    for r in rows:
        try:
            pain = json.loads(r["pain_points_json"] or "[]")
        except json.JSONDecodeError:
            pain = []
        # Pull up to 8 transcript lines as grounding (user + agent only).
        msgs = conn.execute("""
            SELECT role, text FROM messages
            WHERE session_id = ? AND role IN ('user','agent')
              AND text IS NOT NULL AND text != ''
            ORDER BY idx LIMIT 8
        """, (r["session_id"],)).fetchall()
        excerpt_lines = [f"{m['role'].upper()}: {(m['text'] or '')[:200]}" for m in msgs]
        out.append({
            "session_id":   r["session_id"],
            "category":     r["category"],
            "severity":     r["severity"],
            "pain_points":  pain,
            "suggestion":   r["suggestion"],
            "duration_s":   r["duration_seconds"],
            "first_msg":    r["first_user_message"],
            "transcript_excerpt": "\n".join(excerpt_lines),
        })
    return out


def fetch_literal_quote(conn: sqlite3.Connection, session_id: str) -> str | None:
    """Return the first substantive user utterance from the session (>= 3 words)."""
    row = conn.execute("""
        SELECT text FROM messages
        WHERE session_id = ? AND role = 'user'
          AND text IS NOT NULL AND LENGTH(text) > 15
        ORDER BY idx LIMIT 1
    """, (session_id,)).fetchone()
    if not row:
        return None
    return (row["text"] or "").strip()[:250]


# ---------------------------------------------------------------------------
# Clustering via Sonnet
# ---------------------------------------------------------------------------

def cluster_issues(items: list[dict]) -> list[dict]:
    if not config.ANTHROPIC_API_KEY:
        raise EnvironmentError("ANTHROPIC_API_KEY not set")
    client = Anthropic(api_key=config.ANTHROPIC_API_KEY)

    # Compact the items a bit so the prompt fits nicely.
    payload = [
        {
            "session_id":          it["session_id"],
            "category":            it["category"],
            "severity":            it["severity"],
            "pain_points":         it["pain_points"],
            "suggestion":          it["suggestion"],
            "transcript_excerpt":  it["transcript_excerpt"],
        }
        for it in items
    ]

    user_message = (
        "Here are the input classifications to cluster:\n\n"
        + json.dumps(payload, ensure_ascii=False, indent=1)
    )

    resp = client.messages.create(
        model=MODEL,
        max_tokens=16000,
        system=[{"type": "text", "text": CLUSTER_PROMPT,
                 "cache_control": {"type": "ephemeral"}}],
        messages=[{"role": "user", "content": user_message}],
    )
    raw = "".join(b.text for b in resp.content if getattr(b, "type", "") == "text").strip()
    # Always save raw for debugging, pre-strip.
    debug_path = OUT_PATH.parent / "issue_log_raw.json"
    debug_path.parent.mkdir(parents=True, exist_ok=True)
    debug_path.write_text(raw, encoding="utf-8")
    logger.info("Saved raw Sonnet output to %s", debug_path)

    if raw.startswith("```"):
        raw = raw.split("```", 2)[1]
        if raw.startswith("json"):
            raw = raw[4:]
        raw = raw.strip("` \n")

    usage = resp.usage
    logger.info("Sonnet usage — in:%d out:%d cache_read:%d cache_write:%d",
                usage.input_tokens, usage.output_tokens,
                getattr(usage, "cache_read_input_tokens", 0),
                getattr(usage, "cache_creation_input_tokens", 0))

    try:
        data = json.loads(raw)
    except json.JSONDecodeError as e:
        logger.error("Could not parse Sonnet output: %s", e)
        logger.error("Raw (first 2000c): %s", raw[:2000])
        raise

    return data.get("issues") or []


# ---------------------------------------------------------------------------
# Excel writing
# ---------------------------------------------------------------------------

HEADERS = [
    "JOURNEY", "ISSUE", "DESCRIPTION", "# IMPACTED", "PROPOSED OUTCOME",
    "METRIC (VOLUME)", "RAISED BY", "DATE LOGGED", "SEVERITY",
    "REFERENCE", "STATUS",
]

SEV_FILL = {
    "Critical": PatternFill("solid", fgColor="FFC7CE"),
    "High":     PatternFill("solid", fgColor="FFEB9C"),
    "Medium":   PatternFill("solid", fgColor="FFF2CC"),
    "Low":      PatternFill("solid", fgColor="E2EFDA"),
}
HEADER_FILL = PatternFill("solid", fgColor="4A4A4A")
STATUS_FILL = PatternFill("solid", fgColor="C00000")


def write_xlsx(issues: list[dict], conn: sqlite3.Connection, path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Issues"

    # Header row
    for col, name in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = STATUS_FILL if name == "STATUS" else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    today = datetime.now().strftime("%d-%b")
    for i, issue in enumerate(issues, 1):
        row_idx = i + 1
        sev = (issue.get("severity") or "").title()

        # Enrich DESCRIPTION with a literal quote from the first reference session
        refs = issue.get("reference_sessions") or []
        description = issue.get("description") or ""
        quote = None
        if refs:
            quote = fetch_literal_quote(conn, refs[0])
        if quote and f'"{quote}"' not in description:
            description = f'{description}\n\nLiteral quote: "{quote}"'

        # Build REFERENCE column (session IDs, comma-separated)
        ref_str = "\n".join(refs[:3])

        values = [
            issue.get("journey") or "",
            f"{i}. {issue.get('issue_title') or ''}",
            description,
            issue.get("impacted_count"),
            issue.get("proposed_outcome") or "",
            "",  # METRIC (VOLUME) — fill manually
            "George Aguilar",
            today,
            sev,
            ref_str,
            "To Do",
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if col == 9 and sev in SEV_FILL:  # SEVERITY col
                cell.fill = SEV_FILL[sev]

    # Column widths
    widths = {"A": 16, "B": 40, "C": 60, "D": 10, "E": 60, "F": 18,
              "G": 16, "H": 11, "I": 12, "J": 30, "K": 10}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Row heights
    for row in range(2, len(issues) + 2):
        ws.row_dimensions[row].height = 180

    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    conn = sqlite3.connect(config.DB_PATH)
    conn.row_factory = sqlite3.Row

    items = gather_classifications(conn)
    logger.info("Gathered %d high/critical/medium classifications", len(items))

    if not items:
        logger.error("No classifications to cluster.")
        return 1

    issues = cluster_issues(items)
    logger.info("Sonnet returned %d issue clusters", len(issues))

    write_xlsx(issues, conn, OUT_PATH)
    logger.info("Wrote %s (%.1f KB)", OUT_PATH, OUT_PATH.stat().st_size / 1024)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
